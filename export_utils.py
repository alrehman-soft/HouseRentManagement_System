import openpyxl
from openpyxl.styles import Font, Alignment
from tkinter import messagebox
from tkinter import filedialog

def export_to_excel(tree, filename):
    """Export Treeview data to Excel (.xlsx)"""

    # Ask user where to save
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=f"{filename}.xlsx"
    )
    if not file_path:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Write headers
    headers = [tree.heading(col)["text"] for col in tree["columns"]]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_num, child in enumerate(tree.get_children(), start=2):
        values = tree.item(child)["values"]
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)

            if col_num != 2:
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    wb.save(file_path)
    messagebox.showinfo("Success", f"Report exported to {file_path}")
