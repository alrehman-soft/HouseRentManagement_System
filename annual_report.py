import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from create_database import get_connection
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
import print_utils

# ==================== OPEN ANNUAL REPORT ====================
def open_annual_report(parent):
    """Annual Report Tab - Building Wise Annual Summary with styled Treeview"""
    
    # Clear parent frame
    for widget in parent.winfo_children():
        widget.destroy()
    
    main_frame = tk.Frame(parent, bg="#f8fafc")
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # ========== HURAIRA ENTERPRISES TITLE ==========
    title_frame = tk.Frame(main_frame, bg="#1e3a8a",height=70)
    title_frame.pack(fill="x", pady=(0, 10))
    title_frame.pack_propagate(False)
    
    tk.Label(title_frame, 
             text="ABU HURAIRA ENTERPRISES \nAnnual Report", 
             font=("Segoe UI", 18, "bold"), 
             bg="#1e3a8a", fg="white").pack(expand=True)

    # ===== Controls =====
    control_frame = tk.LabelFrame(main_frame, text="Report Controls", 
                                  font=("Segoe UI", 11, "bold"),
                                  bg="#f8fafc", padx=10, pady=10)
    control_frame.pack(fill="x", pady=(0, 10))

    # Year
    tk.Label(control_frame, text="Select Year:", bg="#f8fafc", font=("Segoe UI", 10)).grid(row=0, column=0, padx=5, pady=5, sticky="w")
    year_var = tk.StringVar()
    current_year = datetime.now().year
    years = [str(i) for i in range(current_year, current_year+100)]
    year_combo = ttk.Combobox(control_frame, textvariable=year_var, values=years, state="readonly", width=12)
    year_combo.grid(row=0, column=1, padx=5, pady=5)
    year_combo.set(str(current_year))

    # Building
    tk.Label(control_frame, text="Building:", bg="#f8fafc", font=("Segoe UI", 10)).grid(row=0, column=2, padx=20, pady=5, sticky="w")
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT building_name FROM tenants WHERE building_name IS NOT NULL AND building_name != '' ORDER BY building_name")
    buildings = [row[0] for row in c.fetchall()]
    conn.close()
    
    building_list = ["All Buildings"] + buildings
    building_var = tk.StringVar()
    building_combo = ttk.Combobox(control_frame, textvariable=building_var, values=building_list, state="readonly", width=20)
    building_combo.grid(row=0, column=3, padx=5, pady=5)
    building_combo.current(0)

    # ===== Treeview =====
    report_frame = tk.LabelFrame(main_frame, text="Annual Report", font=("Segoe UI", 11, "bold"),
                                 bg="#f8fafc", padx=10, pady=10)
    report_frame.pack(fill="both", expand=True)

    columns = ("Month","Tenants","Collected","Pending","Rate")
    tree = ttk.Treeview(report_frame, columns=columns, show="headings", height=18)

    # ===== Scrollbar =====
    scrollbar = ttk.Scrollbar(report_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    tree.pack(fill="both", expand=True, side="left")

    style = ttk.Style()
    style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
    style.configure("Treeview", font=("Segoe UI", 10), rowheight=25)
    style.map("Treeview", background=[("selected", "#0c4a6e")], foreground=[("selected", "white")])
    
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center")
    
    tree.column("Month", width=120, anchor="w")
    tree.column("Tenants", width=80, anchor="center")
    for col in ["Collected","Pending","Rate"]:
        tree.column(col, width=120, anchor="e")

    # Tag configuration for styling rows
    tree.tag_configure("oddrow", background="#ffffff")
    tree.tag_configure("evenrow", background="#f1f5f9")
    tree.tag_configure("boldrow", font=("Segoe UI", 10, "bold"), background="#e2e8f0")

    # ===== Buttons =====
    tk.Button(control_frame, text="📊 Generate Report", 
              command=lambda: generate_annual_report(tree, year_var.get(), building_var.get()),
              bg="#1e3a8a", fg="white", font=("Segoe UI", 10, "bold"), width=15).grid(row=0, column=4, padx=20, pady=5)
    
    tk.Button(control_frame, text="🖨️ Print", 
              command=lambda: print_utils.send_to_printer(tree, 
                  title=f"Huraira Enterprises - Annual Report {year_var.get()}", 
                  orientation="landscape"),
              bg="#f59e0b", fg="white", font=("Segoe UI", 10, "bold"), width=12).grid(row=0, column=5, padx=5, pady=5)
    
    tk.Button(control_frame, text="📥 Export", 
              command=lambda: export_report(tree, f"Annual_Report_{year_var.get()}"),
              bg="#10b981", fg="white", font=("Segoe UI", 10, "bold"), width=12).grid(row=0, column=6, padx=5, pady=5)
    return tree

# ==================== GENERATE ANNUAL REPORT ====================
def generate_annual_report(tree, year, building):
    for row in tree.get_children():
        tree.delete(row)
    
    conn = get_connection()
    c = conn.cursor()
    
    # ===== STEP 1: Months list =====
    all_months = ["January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
    
    # ===== STEP 2: Find 1st payment Month =====
    c.execute("""
        SELECT MIN(month_year) FROM payments 
        WHERE month_year LIKE ?
    """, (f"%{year}",))
    
    first_payment = c.fetchone()[0]
    
    if first_payment:
        # 1st payment month
        first_month = first_payment.split('-')[0]
        start_index = all_months.index(first_month)
        print(f"✅ First payment was in: {first_month} (index {start_index})")
    else:
        # If payment not exists
        current_month_num = datetime.now().month
        start_index = current_month_num - 1
        print(f"No payments found, starting from current month: {all_months[start_index]}")
    
    # ===== STEP 3: Current date =====
    current_date = datetime.now()
    current_month_num = current_date.month
    current_year = current_date.year
    
    # ===== STEP 4: Decide months to show =====
    months_to_show = []
    
    if int(year) < current_year:
        # Past year
        months_to_show = all_months[start_index:]
        
    elif int(year) == current_year:
        # Current year
        months_to_show = all_months[start_index:current_month_num]
        
    else:
        # Future year
        messagebox.showinfo("Info", f"Year {year} is in future. No data available.")
        conn.close()
        return
    
    print(f" Showing months: {months_to_show}")
    
    if not months_to_show:
        messagebox.showinfo("Info", "No months to display based on payment history.")
        conn.close()
        return
    
    # ===== STEP 5: Get buildings =====
    if building == "All Buildings":
        c.execute("SELECT DISTINCT building_name FROM tenants WHERE building_name IS NOT NULL AND building_name != '' ORDER BY building_name")
        buildings = [row[0] for row in c.fetchall()]
    else:
        buildings = [building]
    
    grand_total_collected = 0
    grand_total_pending = 0
    row_count = 0
    
    # ===== STEP 6: Process each building =====
    for bld in buildings:
        # Building header
        if building == "All Buildings" and buildings:
            tree.insert("", "end", values=(f"--- {bld} ---", "", "", "", ""), tags=("boldrow",))
            row_count += 1
        
        # Get tenants
        c.execute('''SELECT id, name, flat_no, rent_amount 
                     FROM tenants 
                     WHERE building_name = ? AND status = 'Active'
                     ORDER BY flat_no''', (bld,))
        tenants = c.fetchall()
        if not tenants:
            tenants = [(0, "", "", 0)]
        
        monthly_data = {month: {'collected': 0, 'pending': 0, 'count': 0} for month in months_to_show}
        
        # Collect data
        for tenant in tenants:
            tenant_id, name, flat_no, rent = tenant
            for month in months_to_show:
                month_year = f"{month}-{year}"
                c.execute('''SELECT paid_amount, balance_amount 
                             FROM payments 
                             WHERE tenant_id = ? AND month_year = ?''', 
                          (tenant_id, month_year))
                payment = c.fetchone()
                
                monthly_data[month]['count'] += 1
                if payment:
                    paid, balance = payment
                    monthly_data[month]['collected'] += paid
                    monthly_data[month]['pending'] += balance
                else:
                    # show pending for current or past month
                    monthly_data[month]['pending'] += rent
        
        # Insert rows
        building_total_collected = 0
        for month in months_to_show:
            data = monthly_data[month]
            total_for_month = data['collected'] + data['pending']
            rate = (data['collected'] / total_for_month * 100) if total_for_month > 0 else 0
            
            tag = "evenrow" if row_count % 2 == 0 else "oddrow"
            tree.insert("", "end",
                        values=(month,
                                data['count'],
                                f"Rs {data['collected']:,.0f}",
                                f"Rs {data['pending']:,.0f}",
                                f"{rate:.1f}%"),
                        tags=(tag,))
            row_count += 1
            building_total_collected += data['collected']
        
        building_total_pending = sum(monthly_data[m]['pending'] for m in months_to_show)
        building_rate = (building_total_collected / (building_total_collected + building_total_pending) * 100) if (building_total_collected + building_total_pending) > 0 else 0
        
        tree.insert("", "end", values=(f"{bld} Total", "-", 
                                       f"Rs {building_total_collected:,.0f}",
                                       f"Rs {building_total_pending:,.0f}",
                                       f"{building_rate:.1f}%"), tags=("boldrow",))
        row_count += 1
        
        grand_total_collected += building_total_collected
        grand_total_pending += building_total_pending
    
    # Grand total
    if (grand_total_collected + grand_total_pending) > 0:
        grand_rate = (grand_total_collected / (grand_total_collected + grand_total_pending) * 100)
        tree.insert("", "end", values=("Grand Total", "-", 
                                       f"Rs {grand_total_collected:,.0f}",
                                       f"Rs {grand_total_pending:,.0f}",
                                       f"{grand_rate:.1f}%"), tags=("boldrow",))
    else:
        tree.insert("", "end", values=("No Data", "", "", "", ""), tags=("boldrow",))
    
    conn.close()

def export_report(tree, filename):
    """Export report to Excel"""
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=f"{filename}.xlsx"
    )
    if not file_path:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annual Report"

    # Write headers
    headers = [tree.heading(col)["text"] for col in tree["columns"]]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_num, child in enumerate(tree.get_children(), start=2):
        values = tree.item(child)["values"]
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if col_num != 1:
                cell.alignment = Alignment(horizontal="center")

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    wb.save(file_path)
    messagebox.showinfo("Success", f"Report exported to {file_path}")